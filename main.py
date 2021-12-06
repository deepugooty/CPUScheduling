import os
import re
import json
import numpy as np
import pandas as pd
from openpyxl import load_workbook

# Location of the data folder which contains all the log files
data_directory = 'Data'

# Function to filter the requried data from the log file
def filter_data(file_name, list_of_strings):
    nline = 0
    rlist = []
    # 'r' opens the file in read mode only
    with open(file_name, 'r') as read_obj:
        for line in read_obj:
            nline += 1
            # Checks if list_of_strings i.e. (['Logfile', 'Starting time', 'Finishing time', 'CPU Usage'])
            # is present in the log file content
            for string_to_search in list_of_strings:
                if string_to_search in line:
                    # Removes the all other log file contents
                    rlist.append((string_to_search, nline, line.rstrip()))
    # Returns the list of lines that contains the above mentioned list of strings
    return rlist

# Calculates the CPU allocatio based on the inverse proportion formula
def calculate_cpu_allocation(average_runtime, average_cpu, excel_runtime):
    constant = average_runtime * average_cpu
    final_cpu_allocation = constant / excel_runtime
    final_cpu_allocation = round(final_cpu_allocation)
    return final_cpu_allocation

def main():
    list_of_filenames = ['LayerA', 'LayerB', 'LayerC', 'LayerD']
    # First loops through set of log files belonging to 'LayerA' and repeats for other layers as well
    for lfile in list_of_filenames:
        l = 0
        # Increment of the excel cell value to write the output to the excel file
        row_list = ['C2', 'C3', 'C4', 'C5']
        for row in row_list:
            for filename in os.listdir(data_directory):
                if lfile in filename:
                    # Sends the filename and list to filter_data function and returns a list
                    matched_lines = filter_data(data_directory + '/' + filename, ['Logfile', 'Starting time', 'Finishing time', 'CPU Usage'])
                    temp = {}
                    i = 0
                    data = {}
                    for elem in matched_lines:
                        # Cleans the data by removing empty spaces, '|', '\n', and unnecessary strings
                        result = re.sub(' +', ' ', elem[2])
                        result = result.replace('|', '')
                        result = result.replace(r"\n", "\t")
                        result = result.replace('Qwait: 0h:00m:02s', '')
                        # Writing the output to csv file
                        open('Data\output.csv', 'a').write(result + '\n')
                    # Reading the content of the csv file
                    with open('Data\output.csv', 'r') as f:
                        txt = f.readlines()
                        for eachline in txt:
                            # Creating dataframe with key and value data by splitting at the ':'
                            key = eachline.split(':', 1)[0]
                            value = eachline.split(':', 1)[1]
                            data[key] = value
                        temp[i] = data
                        j = json.dumps(temp)
                        df = pd.read_json(j, orient='index')
                        df = df.replace(r'\n', ' ', regex = True)
                        # Extracting the start time and finishing time columns from the dataframe
                        # and converting it into datetime format
                        dfA = df[' Starting time ']
                        data1 = pd.to_datetime(dfA)
                        dfB = df[' Finishing time ']
                        data2 = pd.to_datetime(dfB)
                        data3 = data2 - data1
                        # Calculating the total run time (start time - finishing time) and converting into hours
                        df[' Run_time '] = data3 / np.timedelta64(1, 'h')
                        df[' Run_time '] = round(df[' Run_time '])
                        df = pd.DataFrame(df, columns=[' Run_time ', ' CPU Usage '])
                        # Appending the output to output1.csv file
                        df.to_csv('Data\output1.csv', mode='a', index=False, header=False)
            # Reading output1.csv file line by line
            with open('Data\output1.csv') as f:
                txt = f.readlines()
                rows_of_numbers = [map(float, line.split(',')) for line in txt]
                # Summing up the average values to get a single average cpu and average runtime value for LayerA
                sums = map(sum, zip(*rows_of_numbers))
                averages = [sum_item / len(txt) for sum_item in sums]
                average_runtime = averages[0]
                average_cpu = averages[1]
                workbook = pd.read_excel('Data/New_OPC_setup.xlsx')
                # Getting the run time cell value from the New_OPC_setup.xlsx file based on its location for LayerA
                excel_runtime = workbook['Run time requirement'].iloc[l]
                # Incrementing the 'l' value to access the LayerB cell value in the next iteration and so on.
                l+=1
                # Condition to check if the runtime is in Minutes ('M') or Hours ('H')
                # If it is in Minutes, convert it into hours by dividing it by 60
                if 'M' in excel_runtime:
                    excel_runtime = re.sub('M', '', excel_runtime)
                    excel_runtime = int(excel_runtime)
                    excel_runtime = excel_runtime/60
                else:
                    excel_runtime = re.sub('H', '', excel_runtime)
                    excel_runtime = int(excel_runtime)
                final_result = calculate_cpu_allocation(average_runtime, average_cpu, excel_runtime)
                wb = load_workbook('Data/New_OPC_setup.xlsx')
                ws = wb['Sheet1']
                # Writing back the final CPU allocation value to excel
                ws[row] = final_result
                # Saving the excel file
                wb.save('Data/New_OPC_setup.xlsx')

if __name__ == '__main__':
    main()